"""
scripts/actualizar_obras_gerencias.py
──────────────────────────────────────
Actualiza diariamente la hoja Obras_Gerencias en Loockups.xlsx.

Fuentes:
  1. OBRAS PRONTO.xlsx — descargado de ProntoNet por GestionComp.
     Columnas usadas:
       Numero           → OBRA_PRONTO  (str, zero-pad 8 dígitos)
       Descripcion      → DESCRIPCION_OBRA
       Unidad operativa → COMPENSABLE

  2. asignacion_gerencias.xlsx (opcional) — enviado por el Director.
     Columnas requeridas: OBRA_PRONTO | GERENCIA
     Si no existe, se preservan las gerencias ya asignadas.

Destino:
  - Hoja Obras_Gerencias en Loockups.xlsx (FAT32).
  - Log de novedades en logs/obras_gerencias_<fecha>.log.

Novedades detectadas:
  - Obras nuevas en ProntoNet (GERENCIA = SIN ASIGNAR).
  - Obras numéricas no encontradas en ProntoNet (se marcan).
  - Cambios en COMPENSABLE.
  - Gerencias asignadas por el Director.

Uso:
  python scripts/actualizar_obras_gerencias.py
  python scripts/actualizar_obras_gerencias.py --dry-run
  python scripts/actualizar_obras_gerencias.py \\
      --obras-pronto /ruta/OBRAS.xlsx \\
      --loockups /ruta/Loockups.xlsx
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

# ── Rutas por defecto ────────────────────────────────────────────────────────
_SCRIPT_DIR = Path(__file__).parent
_BASE_DIR = _SCRIPT_DIR.parent  # gestion_comp/

_INPUT_RAW_DIR = _BASE_DIR / "input_raw"
_OBRAS_PRONTO_DEFAULT = _INPUT_RAW_DIR / "OBRAS PRONTO.xlsx"
_LOGS_DIR = _BASE_DIR / "logs"

_FAT32_COMMON = Path("/media/richard/FAT32/report_gerencias/input_raw/common")
_LOOCKUPS_DEFAULT = _FAT32_COMMON / "Loockups.xlsx"
_ASIGNACION_DEFAULT = _FAT32_COMMON / "asignacion_gerencias.xlsx"

HOJA_OBRAS = "Obras_Gerencias"
GERENCIA_SIN_ASIGNAR = "SIN ASIGNAR"
_HOY = date.today().isoformat()

# ── Logger ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


def _configurar_log_archivo() -> None:
    _LOGS_DIR.mkdir(parents=True, exist_ok=True)
    handler = logging.FileHandler(
        _LOGS_DIR / f"obras_gerencias_{_HOY}.log",
        encoding="utf-8",
    )
    handler.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    )
    logging.getLogger().addHandler(handler)


# ── Lectura de fuentes ───────────────────────────────────────────────────────


def _leer_obras_pronto(ruta: Path) -> pd.DataFrame:
    """
    Lee OBRAS PRONTO.xlsx descargado de ProntoNet.
    Retorna DataFrame con OBRA_PRONTO, DESCRIPCION_OBRA, COMPENSABLE.
    """
    if not ruta.exists():
        raise FileNotFoundError(
            f"OBRAS PRONTO.xlsx no encontrado: {ruta}\n"
            "Ejecutar primero GestionComp para descargarlo."
        )
    df = pd.read_excel(ruta)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all")

    # Nombres de columna esperados (ProntoNet puede exportar con acento)
    _COL_NUMERO = next(
        (c for c in df.columns if c.lower() in ("numero", "número")),
        None,
    )
    _COL_DESC = next(
        (c for c in df.columns if c.lower() == "descripcion"),
        None,
    )
    _COL_UNIDAD = next(
        (c for c in df.columns if c.lower() == "unidad operativa"),
        None,
    )
    faltantes = [
        k
        for k, v in {
            "Numero": _COL_NUMERO,
            "Descripcion": _COL_DESC,
            "Unidad operativa": _COL_UNIDAD,
        }.items()
        if v is None
    ]
    if faltantes:
        raise ValueError(
            f"OBRAS PRONTO.xlsx sin columnas requeridas: {faltantes}"
        )

    df = df.rename(
        columns={
            _COL_NUMERO: "OBRA_PRONTO",
            _COL_DESC: "DESCRIPCION_OBRA",
            _COL_UNIDAD: "COMPENSABLE",
        }
    )
    df["OBRA_PRONTO"] = df["OBRA_PRONTO"].astype(str).str.strip().str.zfill(8)
    df["DESCRIPCION_OBRA"] = df["DESCRIPCION_OBRA"].astype(str).str.strip()
    df["COMPENSABLE"] = df["COMPENSABLE"].astype(str).str.strip()
    return df[
        ["OBRA_PRONTO", "DESCRIPCION_OBRA", "COMPENSABLE"]
    ].drop_duplicates(subset="OBRA_PRONTO")


def _leer_obras_gerencias(loockups: Path) -> pd.DataFrame:
    """Lee la hoja Obras_Gerencias del Loockups.xlsx actual."""
    if not loockups.exists():
        raise FileNotFoundError(f"Loockups.xlsx no encontrado: {loockups}")
    df = pd.read_excel(loockups, sheet_name=HOJA_OBRAS)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all")
    df["OBRA_PRONTO"] = df["OBRA_PRONTO"].astype(str).str.strip()
    return df


def _leer_asignacion_gerencias(ruta: Path) -> dict[str, str]:
    """
    Lee asignacion_gerencias.xlsx del Director.
    Columnas requeridas: OBRA_PRONTO | GERENCIA
    Retorna {OBRA_PRONTO: GERENCIA} o {} si no existe / es inválido.
    """
    if not ruta.exists():
        log.info(
            "asignacion_gerencias.xlsx no encontrado "
            "— se preservan gerencias existentes."
        )
        return {}
    try:
        df = pd.read_excel(ruta)
        df.columns = [str(c).strip().upper() for c in df.columns]
        df = df.dropna(how="all")
        if "OBRA_PRONTO" not in df.columns or "GERENCIA" not in df.columns:
            log.warning(
                "asignacion_gerencias.xlsx sin columnas "
                "OBRA_PRONTO/GERENCIA — ignorado."
            )
            return {}
        df["OBRA_PRONTO"] = (
            df["OBRA_PRONTO"].astype(str).str.strip().str.zfill(8)
        )
        df["GERENCIA"] = df["GERENCIA"].astype(str).str.strip()
        df = df[df["GERENCIA"].str.len() > 0]
        return dict(zip(df["OBRA_PRONTO"], df["GERENCIA"]))
    except Exception as exc:
        log.warning("Error leyendo asignacion_gerencias.xlsx: %s", exc)
        return {}


# ── Cálculo de cambios ───────────────────────────────────────────────────────


def _calcular_cambios(
    df_actual: pd.DataFrame,
    df_pronto: pd.DataFrame,
    mapa_gerencias: dict[str, str],
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    """
    Compara el estado actual de Obras_Gerencias con el nuevo catálogo
    de ProntoNet y el mapa de gerencias del Director.

    Retorna:
      - df_nuevo: DataFrame con todas las filas actualizadas.
      - novedades: lista de cambios detectados.
    """
    novedades: list[dict[str, Any]] = []

    idx_actual: dict[str, Any] = {
        row["OBRA_PRONTO"]: row for _, row in df_actual.iterrows()
    }
    idx_pronto: dict[str, Any] = {
        row["OBRA_PRONTO"]: row for _, row in df_pronto.iterrows()
    }

    # Obras nuevas en ProntoNet
    for op in sorted(set(idx_pronto) - set(idx_actual)):
        row_p = idx_pronto[op]
        novedades.append(
            {
                "tipo": "NUEVA",
                "obra_pronto": op,
                "detalle": (
                    f"{row_p['DESCRIPCION_OBRA']} — "
                    f"COMPENSABLE={row_p['COMPENSABLE']} — "
                    "GERENCIA=SIN ASIGNAR"
                ),
            }
        )

    # Obras numéricas que desaparecieron de ProntoNet
    for op in sorted(set(idx_actual) - set(idx_pronto)):
        if op.isnumeric():
            novedades.append(
                {
                    "tipo": "DESACTIVADA",
                    "obra_pronto": op,
                    "detalle": str(idx_actual[op].get("DESCRIPCION_OBRA", "")),
                }
            )

    # Cambios en COMPENSABLE
    for op in sorted(set(idx_pronto) & set(idx_actual)):
        comp_nuevo = str(idx_pronto[op]["COMPENSABLE"])
        comp_viejo = str(idx_actual[op].get("COMPENSABLE", ""))
        if comp_nuevo != comp_viejo:
            novedades.append(
                {
                    "tipo": "COMPENSABLE_CAMBIO",
                    "obra_pronto": op,
                    "detalle": (
                        f"{idx_actual[op].get('DESCRIPCION_OBRA', '')} "
                        f"— {comp_viejo} → {comp_nuevo}"
                    ),
                }
            )

    # Gerencias asignadas por el Director
    for op, ger_nueva in mapa_gerencias.items():
        _row_a: Any = idx_actual.get(op, {})
        ger_vieja = str(_row_a.get("GERENCIA", ""))  # type: ignore[union-attr]
        if ger_nueva != ger_vieja:
            novedades.append(
                {
                    "tipo": "GERENCIA_ASIGNADA",
                    "obra_pronto": op,
                    "detalle": f"'{ger_vieja}' → '{ger_nueva}'",
                }
            )

    # ── Construir DataFrame actualizado ─────────────────────────────────────
    cols_actuales = list(df_actual.columns)
    extra_cols = [
        c
        for c in cols_actuales
        if c
        not in {
            "OBRA_PRONTO",
            "DESCRIPCION_OBRA",
            "GERENCIA",
            "COMPENSABLE",
        }
    ]

    filas: list[dict[str, Any]] = []

    # Obras provenientes de ProntoNet (fuente de verdad)
    for op, row_p in sorted(idx_pronto.items()):
        row_a: Any = idx_actual.get(op)
        ger_dir = mapa_gerencias.get(op)
        if ger_dir:
            gerencia = ger_dir
        elif row_a is not None and pd.notna(row_a.get("GERENCIA")):
            gerencia = str(row_a["GERENCIA"])
        else:
            gerencia = GERENCIA_SIN_ASIGNAR

        fila: dict[str, Any] = {
            "OBRA_PRONTO": op,
            "DESCRIPCION_OBRA": row_p["DESCRIPCION_OBRA"],
            "GERENCIA": gerencia,
            "COMPENSABLE": row_p["COMPENSABLE"],
        }
        for col in extra_cols:
            fila[col] = row_a[col] if row_a is not None else None
        filas.append(fila)

    # Filas especiales del actual (SIN OBRA, SEDE, etc. — no numéricas)
    for op, row_a in idx_actual.items():
        if op in idx_pronto:
            continue
        if not op.isnumeric():
            fila_esp: dict[str, Any] = {
                "OBRA_PRONTO": op,
                "DESCRIPCION_OBRA": str(row_a.get("DESCRIPCION_OBRA", "")),
                "GERENCIA": str(row_a.get("GERENCIA", "")),
                "COMPENSABLE": str(row_a.get("COMPENSABLE", "")),
            }
            for col in extra_cols:
                fila_esp[col] = row_a.get(col)
            filas.append(fila_esp)

    df_nuevo = pd.DataFrame(filas, columns=cols_actuales)
    return df_nuevo, novedades


# ── Escritura ────────────────────────────────────────────────────────────────


def _escribir_obras_gerencias(
    loockups: Path,
    df_nuevo: pd.DataFrame,
    dry_run: bool,
) -> None:
    """
    Reemplaza la hoja Obras_Gerencias en Loockups.xlsx.
    Preserva todas las demás hojas intactas.
    """
    if dry_run:
        log.info(
            "[DRY-RUN] Se escribirían %d filas en %s.",
            len(df_nuevo),
            HOJA_OBRAS,
        )
        return

    wb = openpyxl.load_workbook(loockups)
    if HOJA_OBRAS in list(wb.sheetnames):
        del wb[HOJA_OBRAS]
    ws = wb.create_sheet(HOJA_OBRAS)

    ws.append(list(df_nuevo.columns))
    for _, row in df_nuevo.iterrows():
        ws.append([None if pd.isna(v) else v for v in row])

    wb.save(loockups)
    log.info(
        "Loockups.xlsx actualizado — %d filas en %s.",
        len(df_nuevo),
        HOJA_OBRAS,
    )


# ── Main ─────────────────────────────────────────────────────────────────────


def _log_novedades(novedades: list[dict[str, Any]]) -> None:
    if not novedades:
        log.info("Sin novedades en %s.", HOJA_OBRAS)
        return
    log.info("Novedades detectadas: %d", len(novedades))
    for n in novedades:
        log.info(
            "  [%s] %s — %s",
            n["tipo"],
            n["obra_pronto"],
            n["detalle"],
        )


def run(
    obras_pronto: Path | None = None,
    loockups: Path | None = None,
    asignacion: Path | None = None,
    dry_run: bool = False,
) -> int:
    """
    Punto de entrada programático (sin argparse).
    Llamar desde main.py u otros módulos.
    Retorna 0 si OK, 1 si hubo error bloqueante.
    """
    _obras = obras_pronto or _OBRAS_PRONTO_DEFAULT
    _loock = loockups or _LOOCKUPS_DEFAULT
    _asig = asignacion or _ASIGNACION_DEFAULT

    log.info("=== Actualización %s — %s ===", HOJA_OBRAS, _HOY)

    try:
        df_pronto = _leer_obras_pronto(_obras)
        log.info("OBRAS PRONTO.xlsx: %d obras leídas.", len(df_pronto))
    except (FileNotFoundError, ValueError) as exc:
        log.error("%s", exc)
        return 1

    try:
        df_actual = _leer_obras_gerencias(_loock)
        log.info("%s actual: %d filas.", HOJA_OBRAS, len(df_actual))
    except Exception as exc:
        log.error("Error leyendo Loockups.xlsx: %s", exc)
        return 1

    mapa_gerencias = _leer_asignacion_gerencias(_asig)
    if mapa_gerencias:
        log.info(
            "asignacion_gerencias.xlsx: %d entradas.",
            len(mapa_gerencias),
        )

    df_nuevo, novedades = _calcular_cambios(
        df_actual, df_pronto, mapa_gerencias
    )
    _log_novedades(novedades)

    try:
        _escribir_obras_gerencias(_loock, df_nuevo, dry_run)
    except Exception as exc:
        log.error("Error escribiendo Loockups.xlsx: %s", exc)
        return 1

    log.info("=== Fin actualización ===")
    return 0


def main() -> int:
    _configurar_log_archivo()

    parser = argparse.ArgumentParser(
        description="Actualiza Obras_Gerencias en Loockups.xlsx"
    )
    parser.add_argument(
        "--obras-pronto",
        type=Path,
        default=_OBRAS_PRONTO_DEFAULT,
        metavar="RUTA",
        help="Ruta a OBRAS PRONTO.xlsx (descargado por GestionComp)",
    )
    parser.add_argument(
        "--loockups",
        type=Path,
        default=_LOOCKUPS_DEFAULT,
        metavar="RUTA",
        help="Ruta a Loockups.xlsx en FAT32",
    )
    parser.add_argument(
        "--asignacion",
        type=Path,
        default=_ASIGNACION_DEFAULT,
        metavar="RUTA",
        help=(
            "Excel del Director con columnas OBRA_PRONTO|GERENCIA "
            "(opcional)"
        ),
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Reporta cambios sin escribir el archivo.",
    )
    args = parser.parse_args()

    return run(
        obras_pronto=args.obras_pronto,
        loockups=args.loockups,
        asignacion=args.asignacion,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    sys.exit(main())
